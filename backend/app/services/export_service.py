import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import List
from datetime import datetime

from app.schemas.work_order import WorkOrderResponse, WorkOrderListResponse


def style_header(ws, num_cols: int):
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws, num_cols: int, max_width: int = 50):
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, max_width)


def fmt_dt(dt: datetime | None, tz_label: str = "") -> str:
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S") + (" " + tz_label if tz_label else "")


def export_single_work_order(wo: WorkOrderResponse, tz_label: str = "UTC") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Order"

    headers = [
        "Reference", "Account Number", "Invoice Number", "PO Number", "Dealer ID",
        "Location Name", "Site Contact", "Address", "Address 2", "City", "State", "Zip", "Primary Phone",
        f"Earliest Start ({tz_label})", f"Planned Start ({tz_label})", f"Due Date ({tz_label})",
        "Status", "Confirmation Status", "Site Timezone",
        "Notes",
        "Created At (UTC)", "Updated At (UTC)",
    ]

    expense_headers = ["Expense Type", "Amount", "Description", "Tech Name"]

    all_headers = headers + expense_headers
    for col, header in enumerate(all_headers, 1):
        ws.cell(row=1, column=col, value=header)

    row = 2
    values = [
        wo.reference, wo.account_number, wo.invoice_number, wo.po_number, wo.dealer_id,
        wo.location_name, wo.site_contact, wo.address_line1, wo.address_line2,
        wo.city, wo.state, wo.zip, wo.primary_phone,
        fmt_dt(wo.earliest_start, tz_label), fmt_dt(wo.planned_start, tz_label), fmt_dt(wo.due_date, tz_label),
        wo.status, wo.confirmation_status, wo.site_timezone,
        wo.notes or "",
        fmt_dt(wo.created_at, "UTC"), fmt_dt(wo.updated_at, "UTC"),
    ]

    expenses_data = getattr(wo, 'expenses', None)
    if expenses_data:
        for i, exp in enumerate(expenses_data):
            r = row + i
            if i == 0:
                for col, val in enumerate(values, 1):
                    ws.cell(row=r, column=col, value=val)
            ws.cell(row=r, column=len(headers) + 1, value=exp.expense_type.value if hasattr(exp.expense_type, 'value') else exp.expense_type)
            ws.cell(row=r, column=len(headers) + 2, value=float(exp.amount))
            ws.cell(row=r, column=len(headers) + 3, value=exp.description or "")
            ws.cell(row=r, column=len(headers) + 4, value=exp.tech_name or "")
    else:
        for col, val in enumerate(values, 1):
            ws.cell(row=row, column=col, value=val)

    num_cols = len(all_headers)
    style_header(ws, num_cols)
    auto_width(ws, num_cols)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_multiple_work_orders(orders: List[WorkOrderListResponse], tz_label: str = "UTC") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Orders"

    headers = [
        "Reference", "Account Number", "Invoice Number",
        "Location Name", "City", "State",
        f"Earliest Start ({tz_label})", f"Planned Start ({tz_label})", f"Due Date ({tz_label})",
        "Status", "Confirmation Status", "Site Timezone",
        "Created At (UTC)", "Updated At (UTC)", "Images",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for i, wo in enumerate(orders):
        row = i + 2
        values = [
            wo.reference, wo.account_number, wo.invoice_number,
            wo.location_name, wo.city, wo.state,
            fmt_dt(wo.earliest_start, tz_label), fmt_dt(wo.planned_start, tz_label), fmt_dt(wo.due_date, tz_label),
            wo.status, wo.confirmation_status, wo.site_timezone,
            fmt_dt(wo.created_at, "UTC"), fmt_dt(wo.updated_at, "UTC"), wo.image_count,
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row, column=col, value=val)

    num_cols = len(headers)
    style_header(ws, num_cols)
    auto_width(ws, num_cols)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
